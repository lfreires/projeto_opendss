import matplotlib.pyplot as plt
import openpyxl
from openpyxl.drawing.image import Image

def gerar_grafico(resultados, caminho_imagem="grafico_tensoes.png"):
    barras = list(resultados.keys())
    tensoes = [t[0] for t in resultados.values()]  # pega apenas a fase 1

    plt.figure(figsize=(10, 6))
    plt.plot(barras, tensoes, marker='o', linestyle='-', color='blue')
    plt.title("Tensões por Barramento")
    plt.xlabel("Barramento")
    plt.ylabel("Tensão Fase 1 (pu)")
    plt.grid(True)
    plt.xticks(rotation=45)
    plt.tight_layout()
    plt.savefig(caminho_imagem)
    plt.close()

def inserir_grafico_no_excel(caminho_excel, resultados, caminho_imagem="grafico_tensoes.png"):
    wb = openpyxl.load_workbook(caminho_excel)

    # Cria aba "Resultados" se não existir
    if "Resultados" not in wb.sheetnames:
        ws = wb.create_sheet("Resultados")
    else:
        ws = wb["Resultados"]

    # Apaga conteúdo antigo (opcional)
    ws.delete_rows(1, ws.max_row)

    # Escreve cabeçalho
    ws.append(["Barramento", "Tensão Fase 1 (pu)"])
    for barra, tensoes in resultados.items():
        ws.append([barra, tensoes[0]])

    # Insere o gráfico
    img = Image(caminho_imagem)
    ws.add_image(img, "E5")

    wb.save(caminho_excel)
    print(f"Gráfico e dados salvos em: {caminho_excel}")
